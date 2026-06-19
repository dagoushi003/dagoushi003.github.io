#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# all_etf_spider.py — 全市场ETF数据采集（GitHub Actions 适配版）
# =====================================================================
# 原版: 本地运行 (Windows, D:\00python\)
# 修改版: CI 环境 (GitHub Actions, UTC+8 时区, 自动 Git Push)
# 修改要点:
#   - 文件路径改为 repo 相对路径
#   - datetime.now() → UTC+8 显式处理
#   - Selenium 添加 CI 专用参数 (--disable-dev-shm-usage)
#   - WebDriverWait 超时从 15s 降到 8s（加快速度）
#   - 深交所 API 添加重试机制
#   - 写入后自动 git commit + push
#   - 交易日前置检查使用 UTC+8 日期
# =====================================================================

import os
import re
import sys
import time
import random
import subprocess
import requests
import pandas as pd
from io import BytesIO
from datetime import datetime, date, timedelta
from chinese_calendar import is_workday
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# =====================================================================
# 配置区（按需修改）
# =====================================================================
# CI 环境: repo 根目录下的相对路径
# 本地测试: 可改为绝对路径
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE_PATH = os.path.join(SCRIPT_DIR, "assets", "posts", "全市场数据分析表02.xlsx")

# 测试开关：True=跳过交易日检查，False=仅交易日运行
# 首次部署建议开 True 跑一次验证，确认没问题后改为 False
TEST_MODE = False

# 时区
TIMEZONE = "Asia/Shanghai"

# =====================================================================
# SSE ETF 完整列表（860只）
# =====================================================================
SSE_LIST = [
 "510010", "510020", "510030", "510040", "510050", "510060", "510090", "510100", "510130", "510150",
 "510160", "510170", "510180", "510190", "510200", "510210", "510230", "510270", "510290", "510300",
 "510310", "510320", "510330", "510350", "510360", "510370", "510380", "510390", "510410", "510500",
 "510510", "510530", "510550", "510560", "510570", "510580", "510590", "510600", "510630", "510650",
 "510660", "510670", "510680", "510710", "510720", "510760", "510770", "510800", "510810", "510850",
 "510880", "510900", "510910", "510950", "510980", "510990", "511010", "511020", "511030", "511060",
 "511070", "511090", "511100", "511110", "511120", "511130", "511150", "511160", "511180", "511190",
 "511200", "511220", "511260", "511270", "511360", "511380", "511520", "511580", "512000", "512010",
 "512020", "512030", "512040", "512050", "512060", "512070", "512080", "512090", "512100", "512120",
 "512130", "512140", "512150", "512160", "512170", "512180", "512190", "512200", "512220", "512240",
 "512250", "512260", "512290", "512330", "512360", "512370", "512380", "512390", "512400", "512430",
 "512450", "512460", "512480", "512500", "512510", "512520", "512530", "512550", "512560", "512570",
 "512580", "512600", "512620", "512630", "512640", "512650", "512660", "512670", "512680", "512690",
 "512700", "512710", "512720", "512730", "512750", "512760", "512770", "512800", "512810", "512820",
 "512870", "512880", "512890", "512900", "512910", "512930", "512940", "512950", "512960", "512970",
 "512980", "512990", "513000", "513010", "513020", "513030", "513040", "513050", "513060", "513070",
 "513080", "513090", "513100", "513110", "513120", "513130", "513140", "513150", "513160", "513170",
 "513180", "513190", "513200", "513210", "513220", "513230", "513240", "513260", "513280", "513290",
 "513300", "513310", "513320", "513330", "513350", "513360", "513380", "513390", "513400", "513500",
 "513520", "513530", "513550", "513560", "513580", "513590", "513600", "513620", "513630", "513650",
 "513660", "513690", "513700", "513720", "513730", "513750", "513770", "513780", "513800", "513810",
 "513820", "513830", "513850", "513860", "513870", "513880", "513890", "513900", "513910", "513920",
 "513930", "513950", "513970", "513980", "513990", "515000", "515010", "515020", "515030", "515050",
 "515060", "515070", "515080", "515090", "515100", "515110", "515120", "515130", "515150", "515160",
 "515170", "515180", "515190", "515200", "515210", "515220", "515230", "515250", "515260", "515290",
 "515300", "515310", "515320", "515330", "515350", "515360", "515370", "515380", "515390", "515400",
 "515450", "515460", "515530", "515550", "515560", "515580", "515590", "515600", "515630", "515640",
 "515650", "515660", "515680", "515700", "515710", "515720", "515730", "515750", "515760", "515770",
 "515790", "515800", "515810", "515850", "515860", "515880", "515890", "515900", "515910", "515920",
 "515950", "515960", "515970", "515980", "515990", "516000", "516010", "516020", "516050", "516060",
 "516070", "516080", "516090", "516100", "516110", "516120", "516130", "516150", "516160", "516180",
 "516190", "516200", "516210", "516220", "516230", "516250", "516260", "516270", "516290", "516300",
 "516310", "516320", "516330", "516350", "516360", "516370", "516380", "516390", "516460", "516500",
 "516510", "516520", "516530", "516550", "516560", "516570", "516580", "516590", "516600", "516610",
 "516620", "516630", "516640", "516650", "516660", "516670", "516700", "516710", "516720", "516730",
 "516750", "516760", "516770", "516780", "516790", "516800", "516810", "516820", "516830", "516840",
 "516850", "516860", "516880", "516890", "516900", "516910", "516920", "516930", "516950", "516960",
 "516970", "516980", "517000", "517010", "517030", "517050", "517080", "517090", "517100", "517110",
 "517120", "517130", "517160", "517170", "517180", "517200", "517300", "517330", "517350", "517360",
 "517380", "517390", "517400", "517520", "517550", "517660", "517770", "517800", "517850", "517880",
 "517900", "517950", "517990", "518600", "518660", "518680", "518800", "518850", "518860", "518880",
 "518890", "520500", "520510", "520520", "520530", "520550", "520560", "520570", "520580", "520590",
 "520600", "520610", "520620", "520630", "520650", "520660", "520670", "520680", "520690", "520700",
 "520710", "520720", "520730", "520760", "520770", "520780", "520790", "520810", "520820", "520830",
 "520840", "520850", "520860", "520870", "520880", "520890", "520900", "520910", "520920", "520930",
 "520940", "520950", "520960", "520970", "520980", "520990", "526000", "526010", "526050", "526070",
 "530000", "530050", "530060", "530080", "530100", "530180", "530280", "530300", "530380", "530530",
 "530580", "530680", "530800", "530880", "551000", "551030", "551060", "551300", "551500", "551510",
 "551520", "551550", "551560", "551580", "551800", "551900", "560000", "560010", "560020", "560030",
 "560050", "560060", "560070", "560080", "560090", "560100", "560110", "560120", "560130", "560150",
 "560160", "560170", "560180", "560190", "560210", "560220", "560230", "560250", "560260", "560270",
 "560280", "560290", "560300", "560310", "560320", "560330", "560350", "560360", "560370", "560380",
 "560390", "560400", "560410", "560420", "560450", "560460", "560470", "560480", "560490", "560500",
 "560510", "560520", "560530", "560550", "560560", "560570", "560580", "560590", "560610", "560620",
 "560630", "560650", "560660", "560670", "560680", "560690", "560700", "560710", "560720", "560750",
 "560770", "560780", "560790", "560800", "560810", "560820", "560830", "560850", "560860", "560870",
 "560880", "560900", "560910", "560930", "560950", "560970", "560980", "560990", "561000", "561010",
 "561030", "561050", "561060", "561080", "561090", "561100", "561120", "561130", "561160", "561170",
 "561180", "561190", "561200", "561220", "561230", "561260", "561280", "561300", "561310", "561320",
 "561330", "561350", "561360", "561370", "561380", "561500", "561510", "561550", "561560", "561570",
 "561580", "561590", "561600", "561660", "561680", "561700", "561750", "561760", "561770", "561780",
 "561790", "561800", "561870", "561880", "561900", "561910", "561920", "561930", "561950", "561960",
 "561980", "561990", "562000", "562010", "562030", "562050", "562060", "562070", "562080", "562300",
 "562310", "562320", "562330", "562340", "562350", "562360", "562380", "562390", "562500", "562510",
 "562520", "562530", "562550", "562560", "562570", "562580", "562590", "562600", "562660", "562700",
 "562800", "562810", "562820", "562850", "562860", "562870", "562880", "562890", "562900", "562910",
 "562920", "562930", "562950", "562960", "562970", "562990", "563000", "563010", "563020", "563030",
 "563050", "563060", "563080", "563090", "563150", "563180", "563200", "563210", "563220", "563230",
 "563280", "563300", "563320", "563330", "563350", "563360", "563380", "563390", "563500", "563510",
 "563520", "563530", "563550", "563560", "563570", "563580", "563590", "563600", "563620", "563630",
 "563650", "563660", "563670", "563680", "563690", "563700", "563750", "563760", "563770", "563780",
 "563790", "563800", "563830", "563850", "563860", "563870", "563880", "563890", "563900", "563930",
 "563960", "563980", "563990", "588000", "588010", "588020", "588030", "588040", "588050", "588060",
 "588070", "588080", "588090", "588100", "588110", "588120", "588130", "588140", "588150", "588160",
 "588170", "588180", "588190", "588200", "588210", "588220", "588230", "588240", "588250", "588260",
 "588270", "588280", "588290", "588300", "588310", "588320", "588330", "588350", "588360", "588370",
 "588380", "588390", "588400", "588410", "588420", "588430", "588450", "588460", "588470", "588480",
 "588500", "588510", "588520", "588550", "588660", "588670", "588680", "588690", "588700", "588710",
 "588720", "588730", "588750", "588760", "588770", "588780", "588790", "588800", "588810", "588820",
 "588830", "588840", "588850", "588860", "588870", "588880", "588890", "588900", "588910", "588920",
 "588930", "588940", "588950", "588960", "588980", "588990", "589000", "589010", "589020", "589030",
 "589050", "589060", "589070", "589080", "589090", "589100", "589110", "589120", "589130", "589150",
 "589160", "589170", "589180", "589190", "589200", "589210", "589220", "589230", "589250", "589260",
 "589270", "589280", "589300", "589320", "589330", "589380", "589500", "589520", "589550", "589560",
 "589580", "589600", "589630", "589660", "589680", "589700", "589720", "589770", "589780", "589800",
 "589820", "589850", "589860", "589880", "589890", "589900", "589950", "589960", "589980", "589990",
]

# =====================================================================
# 全市场ETF数据 Sheet 列位置（13列）
# A=日期 B=代码 C=名称 D=当日份额 E=昨日份额 F=当日净值
# G=当日净份额 H=当日市值 I=昨日市值 J=当日市值变化 K=市值变化百分比
# L=份额变化百分比 M=当日净流入(亿)
# =====================================================================
COL_DATE = 1        # A
COL_CODE = 2        # B
COL_NAME = 3        # C
COL_SHARE_TODAY = 4 # D 当日份额(万份)
COL_SHARE_YEST = 5  # E 昨日份额(万份)
COL_NAV = 6         # F 当日净值
COL_NET_SHARE = 7   # G 当日净份额(万份)
COL_MKT_TODAY = 8   # H 当日市值(亿)
COL_MKT_YEST = 9    # I 昨日市值（亿）
COL_MKT_CHANGE = 10 # J 当日市值变化(亿)
COL_MKT_PCT = 11    # K 市值变化百分比
COL_SHARE_PCT = 12  # L 份额变化百分比
COL_NET_INFLOW = 13 # M 当日净流入(亿)

HEADERS_SSE = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://www.sse.com.cn/"
}
HEADERS_SZSE = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
}

# =====================================================================
# 样式常量
# =====================================================================
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUB_HDR_FILL = PatternFill("solid", fgColor="2E75B6")
SUB_HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=12)
UP_FILL = PatternFill("solid", fgColor="E2EFDA")
DOWN_FILL = PatternFill("solid", fgColor="FCE4D6")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT_PAD = Alignment(horizontal="left", vertical="center", indent=1)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# =====================================================================
# 时间工具：获取 UTC+8 的"今天"和"昨天"
# =====================================================================
def get_today_shanghai():
    """返回上海时区的当前日期 (date 对象)"""
    return date.today()  # GitHub Actions 设置了 TZ=Asia/Shanghai

def get_yesterday_shanghai():
    return get_today_shanghai() - timedelta(days=1)

def get_now_shanghai():
    """返回上海时区的当前 datetime"""
    return datetime.now()  # same as above

# =====================================================================
# 交易日判断（使用 UTC+8 日期）
# =====================================================================
def check_trading_day():
    if TEST_MODE:
        print("⚠️ [测试模式] 已跳过交易日检查，强制运行")
        return True
    yesterday = get_yesterday_shanghai()
    if not is_workday(yesterday):
        print(f"⏭️ 昨天 {yesterday} 非A股交易日，退出。")
        return False
    print(f"✅ 昨天 {yesterday} 是交易日，继续执行")
    return True

# =====================================================================
# Selenium driver 工厂（CI 优化版）
# =====================================================================
def create_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")   # CI/Docker 必需
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--log-level=3")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-background-networking")
    options.add_argument("--disable-sync")
    options.add_argument("--disable-translate")
    options.add_experimental_option("excludeSwitches", ["enable-logging", "enable-automation"])
    # 加快页面加载
    prefs = {
        "profile.default_content_setting_values.notifications": 2,
        "profile.managed_default_content_settings.images": 2,  # 不加载图片
    }
    options.add_experimental_option("prefs", prefs)
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(15)
    return driver

# =====================================================================
# 深交所：一次性获取全量数据（带重试）
# =====================================================================
def fetch_szse_all_data(max_retries=3):
    url_template = (
        "https://fund.szse.cn/api/report/ShowReport"
        "?SHOWTYPE=xlsx&CATALOGID=1000_lf&TABKEY=tab1&random={rand}"
    )
    last_error = None

    for attempt in range(1, max_retries + 1):
        try:
            rand = round(random.random(), 16)
            url = url_template.format(rand=rand)
            print(f"  [深度] 尝试 {attempt}/{max_retries}...")
            resp = requests.get(url, headers=HEADERS_SZSE, timeout=30)
            resp.raise_for_status()

            df = pd.read_excel(BytesIO(resp.content))
            df["当前规模(份)"] = pd.to_numeric(
                df["当前规模(份)"].astype(str).str.replace(",", "").str.strip(),
                errors="coerce"
            )
            df["净值"] = pd.to_numeric(df["净值"], errors="coerce")

            result = []
            for _, row in df.iterrows():
                code = str(row.get("基金代码", "")).strip().zfill(6)
                name = str(row.get("基金简称", "")).strip()
                share = row.get("当前规模(份)")
                nav = row.get("净值")
                if len(code) == 6 and pd.notna(share) and pd.notna(nav) and share > 0:
                    share_wan = round(float(share) / 10000, 4)
                    mkt_val = round(share_wan * float(nav) / 10000, 6)
                    result.append({
                        "code": code,
                        "name": name,
                        "share": share_wan,
                        "nav": float(nav),
                        "mkt_val": mkt_val,
                        "market": "SZSE"
                    })

            print(f" ✅ 深交所ETF: 共 {len(result)} 只 (尝试 {attempt} 次成功)")
            return result

        except requests.exceptions.Timeout as e:
            last_error = f"超时: {e}"
            print(f" ⚠️ 尝试 {attempt} 超时：{e}")
        except requests.exceptions.ConnectionError as e:
            last_error = f"连接失败: {e}"
            print(f" ⚠️ 尝试 {attempt} 连接失败：{e}")
        except Exception as e:
            last_error = str(e)
            print(f" ⚠️ 尝试 {attempt} 失败：{e}")

        if attempt < max_retries:
            wait = 5 * attempt  # 渐退: 5s, 10s, 15s
            print(f"  ⏳ 等待 {wait}s 后重试...")
            time.sleep(wait)

    print(f"❌ 深交所数据获取失败（已重试 {max_retries} 次）：{last_error}")
    return []

# =====================================================================
# 上交所：单只 DOM 提取（CI 优化版，减少 wait timeout）
# =====================================================================
def extract_sse_data_dom(driver, code, max_retries=2):
    for attempt in range(1, max_retries + 1):
        try:
            wait = WebDriverWait(driver, 8)  # 从15s降到8s

            # ---- 1. 份额 & 名称 ----
            wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "div.js_fundSize table.table tbody tr")
            ))
            rows = driver.find_elements(
                By.CSS_SELECTOR, "div.js_fundSize table.table tbody tr"
            )
            if not rows:
                raise ValueError("份额表格行为空")
            tds = rows[0].find_elements(By.TAG_NAME, "td")
            if len(tds) < 2:
                raise ValueError("份额表格列数不足")

            name = tds[2].text.strip() if len(tds) >= 3 else ""
            raw_share = tds[-1].text.replace(",", "").strip()
            share_val = float(raw_share)
            if share_val <= 0:
                raise ValueError("份额为0或负数")

            # ---- 2. 净值 ----
            nav = None
            try:
                nav_tds = driver.find_elements(
                    By.CSS_SELECTOR, "div.js_marketIndex td.colData_val"
                )
                if nav_tds:
                    raw_nav = nav_tds[0].text.replace(",", "").strip()
                    if raw_nav:
                        nav = float(raw_nav)
            except Exception:
                pass

            # ---- 3. 市值（万元转亿） ----
            mkt_val = None
            try:
                mkt_tds = driver.find_elements(
                    By.CSS_SELECTOR, "div.js_transactionOverview td.colData_val"
                )
                if mkt_tds:
                    raw_mkt = mkt_tds[0].text.replace(",", "").strip()
                    if raw_mkt:
                        mkt_val = round(float(raw_mkt) / 10000, 6)
            except Exception:
                pass

            return share_val, nav, mkt_val, name

        except Exception as e:
            if attempt < max_retries:
                time.sleep(1.5)
            else:
                print(f" ⚠️ DOM读取失败 {code}: {e}")
                return None, None, None, ""

# =====================================================================
# 上交所：Selenium 批量抓取（速度优化版）
# =====================================================================
def fetch_sse_all_data(code_list):
    driver = create_driver()
    result = []
    failed = []
    total = len(code_list)
    t_start = time.time()
    base_url = "https://www.sse.com.cn/assortment/fund/list/etfinfo/basic/index.shtml?FUNDID="

    print(f" 🚀 上交所 Selenium 开始，共 {total} 只（耗时约 {total//200}~{total//100} 分钟）...")

    try:
        for idx, code in enumerate(code_list, 1):
            try:
                driver.get(f"{base_url}{code}")
                # 减少 scroll 距离（从1800降到800，加快渲染）
                driver.execute_script("window.scrollTo(0, 800);")
                time.sleep(0.3)  # 短暂等待渲染

                share, nav, mkt_val, name = extract_sse_data_dom(driver, code)

                if share and share > 0:
                    result.append({
                        "code": code,
                        "name": name,
                        "share": share,
                        "nav": nav,
                        "mkt_val": mkt_val,
                        "market": "SSE"
                    })
                    nav_str = f"{nav:.4f}" if nav else "无净值"
                    mkt_str = f"{mkt_val:.4f}亿" if mkt_val else "无市值"
                    print(f" [{idx}/{total}] ✅ {code} {name}: 份额={share:.2f}万份 净值={nav_str} 市值={mkt_str}")
                else:
                    failed.append(code)
                    print(f" [{idx}/{total}] ⚠️ {code}: 获取失败（无份额数据）")

                # 每100只输出进度
                if idx % 100 == 0:
                    elapsed = time.time() - t_start
                    rate = idx / elapsed  # 只/秒
                    remaining = (total - idx) / rate if rate > 0 else 0
                    print(f" 📊 进度 {idx}/{total} ({idx/total*100:.0f}%) | "
                          f"速度 {rate:.1f}只/秒 | "
                          f"已用 {elapsed/60:.1f}分 | "
                          f"预计剩余 {remaining/60:.1f}分")

            except Exception as e:
                failed.append(code)
                print(f" [{idx}/{total}] ❌ {code}: {e}")

    finally:
        driver.quit()

    if failed:
        print(f"\n ⚠️ 共 {len(failed)} 只失败 (成功率 {len(result)}/{total} = {len(result)/total*100:.1f}%)")
        # 输出前20个失败代码便于排查
        print(f"  失败样例: {failed[:20]}")
    else:
        print(f"\n ✅ 上交所全部成功: {len(result)}/{total} 只")
    return result

# =====================================================================
# Excel 工具
# =====================================================================
def ensure_excel(path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if os.path.exists(path):
        return
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "全市场ETF数据"
    headers = [
        "日期", "代码", "名称",
        "当日份额(万份)", "昨日份额(万份)", "当日净值",
        "当日净份额(万份)", "当日市值(亿)", "昨日市值（亿）",
        "当日市值变化(亿)", "市值变化百分比", "份额变化百分比",
        "当日净流入(亿)"
    ]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(1, c, h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = CENTER
    for sname in ["全部排行榜", "大市值排行榜", "小市值排行版"]:
        wb.create_sheet(sname)
    wb.save(path)
    print(f" 📄 新建Excel: {path}")

def get_prev_data(ws):
    """读取上一次写入的数据：代码 → {share, mkt_val, row}"""
    code_map = {}
    for row in range(2, ws.max_row + 1):
        code = ws.cell(row, COL_CODE).value
        if code and isinstance(code, (str, int)) and str(code).strip().isdigit():
            code = str(code).strip().zfill(6)
            share = ws.cell(row, COL_SHARE_TODAY).value
            mkt = ws.cell(row, COL_MKT_TODAY).value
            code_map[code] = {
                "share": float(share) if share is not None else None,
                "mkt_val": float(mkt) if mkt is not None else None,
                "row": row
            }
    return code_map

# =====================================================================
# 写入全市场ETF数据 Sheet
# =====================================================================
def write_data_sheet(ws, all_data, prev_map, today_str):
    written = 0
    total_mkt = sse_mkt = szse_mkt = 0.0
    total_chg = sse_chg = szse_chg = 0.0

    for item in all_data:
        code = item["code"]
        name = item["name"]
        share = item["share"]
        nav = item["nav"]
        mkt_val = item["mkt_val"]
        market = item["market"]

        prev = prev_map.get(code, {})
        share_yest = prev.get("share")
        mkt_yest = prev.get("mkt_val")
        row = prev.get("row", ws.max_row + 1)

        net_share = round(share - share_yest, 4) if share_yest is not None else None
        mkt_change = round(mkt_val - mkt_yest, 6) if (mkt_val is not None and mkt_yest is not None) else None

        if mkt_val:
            total_mkt += mkt_val
            if market == "SSE": sse_mkt += mkt_val
            else: szse_mkt += mkt_val
        if mkt_change:
            total_chg += mkt_change
            if market == "SSE": sse_chg += mkt_change
            else: szse_chg += mkt_change

        ws.cell(row, COL_DATE).value = today_str
        ws.cell(row, COL_CODE).value = code
        ws.cell(row, COL_NAME).value = name
        ws.cell(row, COL_SHARE_TODAY).value = share
        ws.cell(row, COL_SHARE_YEST).value = share_yest
        ws.cell(row, COL_NAV).value = nav
        ws.cell(row, COL_NET_SHARE).value = net_share
        ws.cell(row, COL_MKT_TODAY).value = mkt_val
        ws.cell(row, COL_MKT_YEST).value = mkt_yest
        ws.cell(row, COL_MKT_CHANGE).value = mkt_change

        # K 列：市值变化百分比
        j_ref = ws.cell(row, COL_MKT_CHANGE).coordinate
        i_ref = ws.cell(row, COL_MKT_YEST).coordinate
        ws.cell(row, COL_MKT_PCT).value = f'=IFERROR({j_ref}/{i_ref},"")'
        ws.cell(row, COL_MKT_PCT).number_format = "0.00%"

        # L 列：份额变化百分比
        d_ref = ws.cell(row, COL_SHARE_TODAY).coordinate
        e_ref = ws.cell(row, COL_SHARE_YEST).coordinate
        ws.cell(row, COL_SHARE_PCT).value = f'=IFERROR(({d_ref}-{e_ref})/{e_ref},"")'
        ws.cell(row, COL_SHARE_PCT).number_format = "0.00%"

        # M 列：当日净流入(亿) = 当日净值(F) * 当日净份额(G) / 10000
        f_ref = ws.cell(row, COL_NAV).coordinate
        g_ref = ws.cell(row, COL_NET_SHARE).coordinate
        ws.cell(row, COL_NET_INFLOW).value = f'=IFERROR({f_ref}*{g_ref}/10000,"")'
        ws.cell(row, COL_NET_INFLOW).number_format = "#,##0.0000"

        written += 1

    # 汇总行
    sr = ws.max_row + 2
    sse_data = [d for d in all_data if d["market"] == "SSE"]
    szse_data = [d for d in all_data if d["market"] == "SZSE"]
    for r, label, data, mkt, chg in [
        (sr, "【全市场汇总】", all_data, total_mkt, total_chg),
        (sr + 1, " ↳ 上交所", sse_data, sse_mkt, sse_chg),
        (sr + 2, " ↳ 深交所", szse_data, szse_mkt, szse_chg),
    ]:
        ws.cell(r, COL_DATE).value = today_str
        ws.cell(r, COL_CODE).value = label
        ws.cell(r, COL_NAME).value = f"{len(data)}只"
        ws.cell(r, COL_SHARE_TODAY).value = round(sum(d["share"] for d in data), 2)
        ws.cell(r, COL_MKT_TODAY).value = round(mkt, 4)
        ws.cell(r, COL_MKT_CHANGE).value = round(chg, 4)

    print(f" ✅ 数据Sheet写入: {written}只")
    print(f" 📊 全市场总市值: {total_mkt:,.2f}亿 (沪{sse_mkt:,.2f} + 深{szse_mkt:,.2f})")
    print(f" 📊 全市场市值变化: {total_chg:+.2f}亿 (沪{sse_chg:+.2f} + 深{szse_chg:+.2f})")
    return total_mkt, total_chg

# =====================================================================
# 写入排行榜 Sheet
# =====================================================================
def write_ranking_sheet(ws, df, top_n, today_str):
    ws.delete_rows(1, ws.max_row)

    def get_top(col_name, n, ascending=False):
        valid = df[df[col_name].notna()]
        return valid.sort_values(col_name, ascending=ascending).head(n)[
            ["code", "name", col_name]
        ].reset_index(drop=True)

    inc_share_pct = get_top("share_pct", top_n, ascending=False)
    inc_mkt = get_top("mkt_change", top_n, ascending=False)
    inc_mkt_pct = get_top("mkt_pct", top_n, ascending=False)
    dec_share_pct = get_top("share_pct", top_n, ascending=True)
    dec_mkt = get_top("mkt_change", top_n, ascending=True)
    dec_mkt_pct = get_top("mkt_pct", top_n, ascending=True)

    # 标题行
    ws.row_dimensions[1].height = 22
    ws.cell(1, 1, today_str).font = TITLE_FONT
    ws.cell(1, 1).alignment = CENTER

    r_inc_hdr = 2
    r_inc_sub = 3
    r_inc_data = 4

    # 增长区块：表头
    inc_groups = [
        (1, f"份额变化百分比排名（增加前{top_n}）", inc_share_pct, "份额变化百分比", "0.00%"),
        (4, f"市值增长最多排名（前{top_n}）", inc_mkt, "增长值（亿）", "#,##0.0000"),
        (7, f"市值增长百分比排名（前{top_n}）", inc_mkt_pct, "百分比", "0.00%"),
    ]
    for col_start, title, _, val_hdr, _ in inc_groups:
        hcell = ws.cell(r_inc_hdr, col_start, title)
        hcell.font = HDR_FONT; hcell.fill = HDR_FILL; hcell.alignment = CENTER
        ws.merge_cells(start_row=r_inc_hdr, start_column=col_start,
                       end_row=r_inc_hdr, end_column=col_start + 2)
        for c, h in zip(range(col_start, col_start + 3), ["代码", "名称", val_hdr]):
            sc = ws.cell(r_inc_sub, c, h)
            sc.font = SUB_HDR_FONT; sc.fill = SUB_HDR_FILL; sc.alignment = CENTER

    # 增长区块：数据
    for i in range(top_n):
        r = r_inc_data + i
        for col_start, _, src_df, _, fmt in inc_groups:
            if i < len(src_df):
                rd = src_df.iloc[i]
                ws.cell(r, col_start, rd["code"]).alignment = CENTER
                ws.cell(r, col_start + 1, rd["name"]).alignment = LEFT_PAD
                vc = ws.cell(r, col_start + 2, rd.iloc[2])
                vc.number_format = fmt; vc.alignment = RIGHT
                for c in range(col_start, col_start + 3):
                    ws.cell(r, c).fill = UP_FILL

    # 减少区块
    r_dec_start = r_inc_data + top_n + 2
    dec_groups = [
        (1, f"份额变化百分比排名（减少前{top_n}）", dec_share_pct, "份额变化百分比", "0.00%"),
        (4, f"市值减少最多排名（前{top_n}）", dec_mkt, "减少值（亿）", "#,##0.0000"),
        (7, f"市值减少百分比排名（前{top_n}）", dec_mkt_pct, "百分比", "0.00%"),
    ]
    for col_start, title, src_df, val_hdr, fmt in dec_groups:
        hcell = ws.cell(r_dec_start, col_start, title)
        hcell.font = HDR_FONT; hcell.fill = HDR_FILL; hcell.alignment = CENTER
        ws.merge_cells(start_row=r_dec_start, start_column=col_start,
                       end_row=r_dec_start, end_column=col_start + 2)
        for c, h in zip(range(col_start, col_start + 3), ["代码", "名称", val_hdr]):
            sc = ws.cell(r_dec_start + 1, c, h)
            sc.font = SUB_HDR_FONT; sc.fill = SUB_HDR_FILL; sc.alignment = CENTER
        for i in range(min(top_n, len(src_df))):
            r = r_dec_start + 2 + i
            rd = src_df.iloc[i]
            ws.cell(r, col_start, rd["code"]).alignment = CENTER
            ws.cell(r, col_start + 1, rd["name"]).alignment = LEFT_PAD
            vc = ws.cell(r, col_start + 2, rd.iloc[2])
            vc.number_format = fmt; vc.alignment = RIGHT
            for c in range(col_start, col_start + 3):
                ws.cell(r, c).fill = DOWN_FILL

    # 列宽
    for col, width in zip("ABCDEFGHI", [10, 22, 14, 10, 22, 14, 10, 22, 14]):
        ws.column_dimensions[col].width = width

# =====================================================================
# 主写入流程
# =====================================================================
def write_excel(all_data, today_str):
    ensure_excel(EXCEL_FILE_PATH)

    wb = load_workbook(EXCEL_FILE_PATH)
    for sname in ["全市场ETF数据", "全部排行榜", "大市值排行榜", "小市值排行版"]:
        if sname not in wb.sheetnames:
            wb.create_sheet(sname)

    ws_data = wb["全市场ETF数据"]
    if ws_data.max_row == 1 and ws_data.cell(1, 1).value is None:
        headers = [
            "日期", "代码", "名称",
            "当日份额(万份)", "昨日份额(万份)", "当日净值",
            "当日净份额(万份)", "当日市值(亿)", "昨日市值（亿）",
            "当日市值变化(亿)", "市值变化百分比", "份额变化百分比"
        ]
        for c, h in enumerate(headers, 1):
            cell = ws_data.cell(1, c, h)
            cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = CENTER

    prev_map = get_prev_data(ws_data)
    write_data_sheet(ws_data, all_data, prev_map, today_str)

    # ---- 构建排行榜 DataFrame ----
    rows = []
    for item in all_data:
        code = item["code"]
        share = item["share"]
        mkt_val = item["mkt_val"]
        prev = prev_map.get(code, {})
        share_yest = prev.get("share")
        mkt_yest = prev.get("mkt_val")
        mkt_change = round(mkt_val - mkt_yest, 6) if (mkt_val is not None and mkt_yest is not None) else None
        mkt_pct = mkt_change / mkt_yest if (mkt_change is not None and mkt_yest and mkt_yest != 0) else None
        share_pct = (share - share_yest) / share_yest if (share_yest and share_yest != 0) else None
        rows.append({
            "code": code,
            "name": item["name"],
            "share_pct": share_pct,
            "mkt_today": mkt_val,
            "mkt_change": mkt_change,
            "mkt_pct": mkt_pct,
        })

    df_all = pd.DataFrame(rows)
    df_large = df_all[df_all["mkt_today"].notna() & (df_all["mkt_today"] >= 100)]
    df_small = df_all[df_all["mkt_today"].notna() & (df_all["mkt_today"] < 100)]

    write_ranking_sheet(wb["全部排行榜"], df_all, 30, today_str)
    write_ranking_sheet(wb["大市值排行榜"], df_large, 20, today_str)
    write_ranking_sheet(wb["小市值排行版"], df_small, 20, today_str)

    wb.save(EXCEL_FILE_PATH)
    print(f"\n✨ 全部写入完成 → {EXCEL_FILE_PATH}")

# =====================================================================
# Git 提交与推送
# =====================================================================
def git_commit_and_push(repo_dir, today_str):
    """将更新的 Excel 提交并推送到 GitHub"""
    try:
        os.chdir(repo_dir)

        # 检查是否有变更
        result = subprocess.run(
            ["git", "status", "--porcelain", "assets/posts/全市场数据分析表02.xlsx"],
            capture_output=True, text=True, timeout=30
        )
        if not result.stdout.strip():
            print(" ℹ️ 无变更，跳过推送")
            return True

        # git add
        subprocess.run(
            ["git", "add", "assets/posts/全市场数据分析表02.xlsx"],
            check=True, capture_output=True, timeout=30
        )

        # git commit
        commit_msg = f"📊 全市场ETF数据更新 {today_str}"
        subprocess.run(
            ["git", "commit", "-m", commit_msg],
            check=True, capture_output=True, timeout=30
        )

        # git push
        push_result = subprocess.run(
            ["git", "push"],
            capture_output=True, text=True, timeout=60
        )
        if push_result.returncode == 0:
            print(f" ✅ Git 推送成功: {commit_msg}")
        else:
            print(f" ⚠️ Git 推送可能失败: {push_result.stderr[:200]}")
            return False

        return True

    except subprocess.TimeoutExpired:
        print(" ⚠️ Git 操作超时")
        return False
    except subprocess.CalledProcessError as e:
        print(f" ⚠️ Git 操作失败: {e.stderr.decode() if isinstance(e.stderr, bytes) else e.stderr}")
        return False
    except Exception as e:
        print(f" ⚠️ Git 异常: {e}")
        return False

# =====================================================================
# 去重检查：今天的日期在Excel里是否已有数据
# =====================================================================
def check_today_already_done(today_str):
    """检查Excel中是否已有今天日期的数据，有则跳过本次运行"""
    if not os.path.exists(EXCEL_FILE_PATH):
        return False
    try:
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb["全市场ETF数据"]
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, COL_DATE).value
            if val and str(val).strip() == today_str:
                print(f" ℹ️ 今天 {today_str} 的数据已存在，跳过本次运行")
                wb.close()
                return True
        wb.close()
    except Exception as e:
        print(f" ⚠️ 去重检查异常: {e}")
    return False


# =====================================================================
# 主流程
# =====================================================================
def run():
    if not check_trading_day():
        return

    start = get_now_shanghai()
    today_str = start.strftime("%Y-%m-%d")

    # 去重：如果今天数据已有，跳过（用于多定时重试场景）
    if check_today_already_done(today_str):
        return

    sep = "=" * 55

    print(sep)
    print(f" 全市场ETF数据采集 — {today_str} {start.strftime('%H:%M:%S')}")
    print(f" 时区: {TIMEZONE}  |  测试模式: {'开' if TEST_MODE else '关'}")
    print(f" 文件: {EXCEL_FILE_PATH}")
    print(sep)

    # ==================== 深交所 ====================
    print("\n【深交所】获取ETF数据...")
    szse_data = fetch_szse_all_data()

    # ==================== 上交所 ====================
    print("\n【上交所】获取ETF数据...")
    sse_data = fetch_sse_all_data(SSE_LIST)

    # ==================== 合并 ====================
    all_data = szse_data + sse_data
    print(f"\n📦 合并: 深交所{len(szse_data)}只 + 上交所{len(sse_data)}只 = 共{len(all_data)}只")

    if not all_data:
        print("❌ 未获取到任何数据，放弃写入")
        return

    # ==================== 写入 ====================
    print("\n【写入Excel】...")
    write_excel(all_data, today_str)

    # ==================== Git 推送 ====================
    print("\n【Git 推送】...")
    repo_dir = SCRIPT_DIR
    git_commit_and_push(repo_dir, today_str)

    elapsed = (get_now_shanghai() - start).seconds
    print(f"\n⏱️ 总耗时: {elapsed // 60}分{elapsed % 60}秒")
    print(sep)


if __name__ == "__main__":
    run()
